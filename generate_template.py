import openpyxl
from openpyxl.styles import Font

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Student Upload Template"

# Define headers
headers = [
    'first_name', 'last_name', 'roll_number', 'grade', 
    'section', 'date_of_birth', 'gender', 'phone', 'email'
]

# Write headers
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)

# Add a sample row (optional, but helpful)
sample_data = [
    'Rahim', 'Uddin', '101', '10', 
    'A', '2005-01-01', 'Male', '01700000000', 'rahim@example.com'
]

for col_num, data in enumerate(sample_data, 1):
    ws.cell(row=2, column=col_num).value = data

# Save the file
file_path = "D:\\django project\\School management\\static\\samples\\student_upload_template.xlsx"
wb.save(file_path)
print(f"Template saved to {file_path}")
